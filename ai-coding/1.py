# filename: generate_excel.py

import openpyxl
from openpyxl import Workbook

def create_excel_file():
    wb = Workbook()

    # --- Sheet1: 日常运营数据 (Daily_Operation_Data) ---
    ws1 = wb.active
    ws1.title = "Daily_Operation_Data"
    header1 = ["Date", "DAU", "New_Registers", "1-Day_Retention",
               "7-Day_Retention", "Paying_Users", "Revenue(iOS+Android)",
               "iOS_Share", "Android_Share"]
    data1 = [
        ["2025-01-01", 100000, 2500, "58%", "22%", 4800, 5000, "60%", "40%"],
        ["2025-01-02", 102000, 2700, "59%", "23%", 4900, 5200, "62%", "38%"],
        ["2025-01-03", 98000, 2400, "57%", "21%", 4700, 4900, "61%", "39%"],
        ["2025-01-04", 105000, 3000, "61%", "24%", 5000, 5500, "59%", "41%"],
        ["2025-01-05", 110000, 3200, "62%", "25%", 5200, 5800, "60%", "40%"],
        ["2025-01-06", 108000, 3100, "60%", "24%", 5100, 5700, "60%", "40%"],
        ["2025-01-07", 107000, 2800, "59%", "23%", 5050, 5600, "61%", "39%"],
        ["2025-01-08", 112000, 3500, "62%", "25%", 5300, 6000, "62%", "38%"],
        ["2025-01-09", 115000, 3600, "63%", "26%", 5400, 6200, "61%", "39%"],
        ["2025-01-10", 111000, 3100, "60%", "24%", 5200, 5800, "60%", "40%"],
        ["2025-01-11", 109000, 2900, "58%", "23%", 5000, 5500, "60%", "40%"],
        ["2025-01-12", 113000, 3400, "61%", "25%", 5300, 6000, "62%", "38%"],
        ["2025-01-13", 117000, 3700, "64%", "26%", 5600, 6500, "63%", "37%"],
        ["2025-01-14", 120000, 4000, "65%", "27%", 5800, 6800, "64%", "36%"],
    ]

    ws1.append(header1)
    for row in data1:
        ws1.append(row)

    # --- Sheet2: 渠道投放数据 (Channel_Data) ---
    ws2 = wb.create_sheet("Channel_Data")
    header2 = ["Date", "Channel", "Impressions", "Clicks", "CTR",
               "Conversions", "Conversion_Rate", "CPC", "Spend"]
    data2 = [
        ["2025-01-01", "WeChat", 100000, 12000, "12.0%", 3000, "25.0%", 2, 24000],
        ["2025-01-01", "Douyin", 80000, 6400, "8.0%", 1600, "25.0%", 1.5, 9600],
        ["2025-01-01", "Baidu SEM", 50000, 2500, "5.0%", 400, "16.0%", 3, 7500],
        ["2025-01-01", "Offline", 20000, 2000, "10.0%", 300, "15.0%", None, None],
        ["2025-01-02", "WeChat", 110000, 13200, "12.0%", 3300, "25.0%", 2, 26400],
        ["2025-01-02", "Douyin", 85000, 6800, "8.0%", 1700, "25.0%", 1.5, 10200],
        ["2025-01-02", "Baidu SEM", 52000, 2600, "5.0%", 420, "16.2%", 3, 7800],
        ["2025-01-02", "Offline", 21000, 2200, "10.5%", 320, "14.5%", None, None],
        ["2025-01-03", "WeChat", 90000, 10800, "12.0%", 2700, "25.0%", 2, 21600],
        ["2025-01-03", "Douyin", 78000, 6240, "8.0%", 1560, "25.0%", 1.5, 9360],
        ["2025-01-03", "Baidu SEM", 49000, 2450, "5.0%", 390, "15.9%", 3, 7350],
        ["2025-01-03", "Offline", 19000, 1950, "10.3%", 280, "14.4%", None, None],
        ["2025-01-04", "WeChat", 120000, 14400, "12.0%", 3600, "25.0%", 2, 28800],
        ["2025-01-04", "Douyin", 88000, 7040, "8.0%", 1760, "25.0%", 1.5, 10560],
        ["2025-01-04", "Baidu SEM", 55000, 2750, "5.0%", 440, "16.0%", 3, 8250],
        ["2025-01-04", "Offline", 23000, 2300, "10.0%", 350, "15.2%", None, None],
    ]
    ws2.append(header2)
    for row in data2:
        ws2.append(row)

    # --- Sheet3: 漏斗数据 (Funnel_Data) ---
    ws3 = wb.create_sheet("Funnel_Data")
    header3 = ["Date", "HomePage_Visits", "ProductPage_Visits",
               "AddToCart", "CheckoutClicks", "FinalPurchases"]
    data3 = [
        ["2025-01-01", 50000, 40000, 20000, 10000, 8000],
        ["2025-01-02", 52000, 42000, 21000, 10800, 8500],
        ["2025-01-03", 48000, 38000, 19000, 9500, 7600],
        ["2025-01-04", 55000, 45000, 22000, 11500, 9000],
        ["2025-01-05", 58000, 47000, 23000, 12000, 9200],
        ["2025-01-06", 56000, 46000, 22500, 11500, 8800],
        ["2025-01-07", 57000, 46500, 22800, 11700, 9000],
        ["2025-01-08", 60000, 49000, 24500, 12800, 9800],
        ["2025-01-09", 62000, 50500, 25000, 13000, 10000],
        ["2025-01-10", 59000, 48000, 24000, 12500, 9500],
        ["2025-01-11", 58000, 47000, 23500, 12000, 9200],
        ["2025-01-12", 60500, 49500, 24800, 12700, 9700],
        ["2025-01-13", 63000, 52000, 26000, 13500, 10500],
        ["2025-01-14", 65000, 54000, 27000, 14000, 11000],
    ]
    ws3.append(header3)
    for row in data3:
        ws3.append(row)

    # --- Sheet4: 地域分布 (Region_Data) ---
    ws4 = wb.create_sheet("Region_Data")
    header4 = ["Date", "Region", "Daily_Active_Users", "Daily_Purchases"]
    data4 = [
        ["2025-01-01", "Guangdong", 15000, 1200],
        ["2025-01-01", "Jiangsu", 10000, 800],
        ["2025-01-01", "Zhejiang", 8000, 600],
        ["2025-01-01", "Beijing", 7500, 550],
        ["2025-01-01", "Shanghai", 6500, 500],
        ["2025-01-01", "Others", 53000, 3150],

        ["2025-01-02", "Guangdong", 15500, 1250],
        ["2025-01-02", "Jiangsu", 10200, 810],
        ["2025-01-02", "Zhejiang", 8300, 620],
        ["2025-01-02", "Beijing", 7600, 560],
        ["2025-01-02", "Shanghai", 6600, 510],
        ["2025-01-02", "Others", 53600, 3200],

        ["2025-01-03", "Guangdong", 14800, 1150],
        ["2025-01-03", "Jiangsu", 9800, 780],
        ["2025-01-03", "Zhejiang", 7900, 580],
        ["2025-01-03", "Beijing", 7400, 540],
        ["2025-01-03", "Shanghai", 6400, 490],
        ["2025-01-03", "Others", 51200, 3060],
    ]
    ws4.append(header4)
    for row in data4:
        ws4.append(row)

    # --- Sheet5: 客服反馈 (Feedback_Data) ---
    ws5 = wb.create_sheet("Feedback_Data")
    header5 = ["Ticket_ID", "Date", "User_ID", "Category",
               "Response_Time(minutes)", "Satisfaction(1-5)", "Comments"]
    data5 = [
        ["TK001", "2025-01-01", "USER1001", "Payment Issue", 3, 5, "Payment failed once, but resolved quickly."],
        ["TK002", "2025-01-01", "USER1020", "Account Problem", 2, 4, "Couldn't log in. Service was good."],
        ["TK003", "2025-01-01", "USER1105", "Order Inquiry", 4, 3, "Delayed response. Eventually solved."],
        ["TK004", "2025-01-01", "USER1110", "General Consultation", 2, 5, "Fast and helpful."],
        ["TK005", "2025-01-02", "USER1203", "Payment Issue", 5, 3, "Issue took some time to fix."],
        ["TK006", "2025-01-02", "USER1307", "Order Inquiry", 3, 4, "Wanted to check shipping progress."],
        ["TK007", "2025-01-02", "USER1322", "Product Content", 2, 5, "Info missing in product description."],
        ["TK008", "2025-01-03", "USER1401", "Account Problem", 3, 4, "Reset password link not working."],
        ["TK009", "2025-01-03", "USER1450", "Payment Issue", 4, 3, "Repeated transaction error."],
        ["TK010", "2025-01-03", "USER1555", "General Consultation", 1, 5, "Quick and friendly support."],
        ["TK011", "2025-01-04", "USER1602", "Order Inquiry", 2, 4, "Asked about delivery date."],
        ["TK012", "2025-01-04", "USER1708", "Product Content", 2, 5, "Needed more color options."],
        ["TK013", "2025-01-05", "USER1801", "Payment Issue", 6, 2, "Slow resolution. Unhappy user."],
        ["TK014", "2025-01-05", "USER1810", "Account Problem", 3, 4, "Locked out of account temporarily."],
        ["TK015", "2025-01-05", "USER1825", "General Consultation", 2, 5, "Asked about membership benefits."],
    ]
    ws5.append(header5)
    for row in data5:
        ws5.append(row)

    # 保存 Excel
    wb.save("SampleData.xlsx")


if __name__ == "__main__":
    create_excel_file()
    print("Excel 文件已生成：SampleData.xlsx")
